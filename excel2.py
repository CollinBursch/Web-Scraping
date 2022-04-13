import openpyxl as xl

wb = xl.load_workbook('2_example.xlsx')

sm = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1.value)

print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)

for i in range(1,sheet1.max_row + 1):
    print(sheet1.cell(i,2).value)

print(sheet1.max_row)
print(sheet1.max_column)

#convert letter to a nuber for columns
print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

#convert back to a number, in case you want to iterate through columns 
print(xl.utils.column_index_form_string('AHP'))

for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)
        print('--- END OF COLUMNS ---')
    
    print()
    print('--- END OF ROW ---')
    print()

    for currentrow in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, max_column=sheet1.max_column):
        print(currentrow[0].value)
        print(currentrow[1].value)
        print(currentrow[2].value)