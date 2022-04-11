
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

movie_table = soup.find('table')
#print(movie_table)

movie_rows = movie_table.findAll('tr')
#print(movie_rows[1])

wb = xl.Workbook()

ws = wb.active

ws.title = 'BoxOfficeReport'

#write headings
ws['A1'] = 'No. '
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = x + 'of Total Gross'

for x in range(1, 6):
    td = movie_rows[x].findAll('td')
    ranking = td[0].text
    title = td[1].text
    gross = int(td[5].text .replace(",","").replace("$",""))
    total_gross = int(td[7].text .replace(",","").replace("$",""))
    release_date = td[8].text

    percent_gross = round((gross/total_gross)*100,2)

    ws['A' + str(x+1)] = ranking
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['D' + str(x+1)] = gross
    ws['E' + str(x+1)] = total_gross
    ws['F' + str(x+1)] = str(percent_gross) + '%'

#adjust column size
ws.column_dimensions['A'].width = 5
ws.column_dimensions['A'].width = 30
ws.column_dimensions['A'].width = 25
ws.column_dimensions['A'].width = 16
ws.column_dimensions['A'].width = 20
ws.column_dimensions['A'].width = 26

#apply fromatting to the header row
header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save('BoxOfficeReport.xlsx')
