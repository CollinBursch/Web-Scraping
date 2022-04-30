from symtable import Symbol
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from twilio.rest import Client

def twilio_text(crypto_name, crypto_symbol, crypto_price):
    accountSID = 'ACa8edc29a571e8e41969bcf6e35aee273'
    AuthToken = '1b51ad1b9327bcdfdc3997122ac36fea'
    client = Client(accountSID, AuthToken)
    TwilioNumber = '+14843107924'
    mycellphone = '+14154196778'
    message_string = crypto_name + ' (' + crypto_symbol + ') ' + 'values have dropped below ' + str(crypto_price)
    textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber, body=message_string)
    print(textmessage.status)

def main():
    url = 'https://cryptoslate.com/coins/'
    req = Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'})
    webpage = urlopen(req).read()		

    soup = BeautifulSoup(webpage, 'html.parser')

    title = soup.title

    crypto_table = soup.find('table')
    crypto_rows = crypto_table.findAll('tr')

    wb = xl.Workbook()

    ws = wb.active

    ws.title = 'TopCryptoReport'


    ws['A1'] = 'Name'
    ws['B1'] = 'Symbol'
    ws['C1'] = 'Price'
    ws['D1'] = '24 Hour Percent Change'
    ws['E1'] = '24 Hour Value Change'

    for x in range(1, 6):
        td = crypto_rows[x].findAll('td')
        name = td[2].find('a').text.split()
        symbol = name[-1]
        name = ' '.join(name[:-1])
        price = float(td[4].text.replace(',','').replace('$',''))
        percent_change = float(td[5].find('span').text.strip('%'))
        old_value = round(price / (1 + (percent_change/100)), 2)
        value_change = round((old_value * (percent_change/100)),2)
        print(name, symbol, price, percent_change, value_change)
        if (symbol == 'BTC' and price < 40000) or (symbol == 'ETH' and price < 3000.00):
            twilio_text(name, symbol, price)
        


        ws['A' + str(x+1)] = name
        ws['B' + str(x+1)] = symbol
        ws['C' + str(x+1)] = price
        ws['D' + str(x+1)] = str(percent_change) + '%'
        ws['E' + str(x+1)] = value_change

        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

    wb.save('TopCryptoReport.xlsx')


main()



