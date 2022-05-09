import openpyxl as xl

# Prompt:
# - Make an excel file that writes the x and y values of the function f(x) = x^2 + 1
# - Do this for values of X from 1 to 100

def main():
    # Making a new excel file stuff
    wb = xl.Workbook() # Make new file object
    ws = wb.active # Activate new sheet
    ws.title = 'MathPoints' # Title of sheet in new file

    # Adding header row (letter = Column, number = row)
    ws['A1'] = 'X - input'
    ws['B1'] = 'Y - output'

    for x in range(1, 101):
        # Adding values to columns
        ws['A' + str(x + 1)] = x
        ws['B' + str(x + 1)] = x**2 + 1

    # Basic styling
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    # Write file to system
    wb.save('Math.xlsx')

main()